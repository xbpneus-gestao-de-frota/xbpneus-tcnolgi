#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo de Análise de Imagens e Preenchimento de Formulário de Garantia
Integração com WhatsApp para recebimento de fotos
"""

import os
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from ..settings import IA_CONFIG
from .ia_analise_pneus import IAnalisePneus

class AnalisadorImagemGarantia:
    """
    Sistema de análise de imagens de pneus e preenchimento automático
    do formulário de garantia
    """
    
    def __init__(self, template_path: Optional[str] = None):
        """
        Inicializa o analisador
        
        Args:
            template_path: Caminho para o template do formulário Excel
        """
        default_template_dir = Path(IA_CONFIG.get('MODEL_PATH'))
        default_template = default_template_dir / 'FORMULARIOPARAGARANTIA.xlsx'
        if template_path:
            self.template_path = Path(template_path)
        else:
            self.template_path = default_template
        self.ia = IAnalisePneus()
        
        # Mapeamento das seções do formulário
        self.secoes = {
            'A': {'linha': 3, 'titulo': 'Consumidor'},
            'B': {'linha': 7, 'titulo': 'Distribuidor/Revendedor'},
            'C': {'linha': 11, 'titulo': 'Produto Reclamado'},
            'D': {'linha': 15, 'titulo': 'Veículo Utilizado'},
            'E': {'linha': 18, 'titulo': 'Condição de Uso'},
            'F': {'linha': 21, 'titulo': 'Defeito Alegado'},
            'G': {'linha': 26, 'titulo': 'Resultado da Análise'},
            'I': {'linha': 45, 'titulo': 'Dados do Examinador'}
        }
    
    def analisar_imagens_whatsapp(self, imagens: List[str], contexto: Dict) -> Dict:
        """
        Analisa imagens recebidas via WhatsApp
        
        Args:
            imagens: Lista de caminhos para as imagens
            contexto: Contexto adicional (mensagens do usuário, etc)
        
        Returns:
            Análise completa das imagens
        """
        resultado = {
            'timestamp': datetime.now().isoformat(),
            'total_imagens': len(imagens),
            'analises_individuais': [],
            'diagnostico_consolidado': {},
            'recomendacao': '',
            'nivel_urgencia': 'BAIXA'
        }
        
        for idx, img_path in enumerate(imagens, 1):
            analise_img = self._analisar_imagem_individual(img_path, idx)
            resultado['analises_individuais'].append(analise_img)
        
        # Consolidar diagnóstico
        resultado['diagnostico_consolidado'] = self._consolidar_diagnostico(
            resultado['analises_individuais'],
            contexto
        )
        
        # Determinar nível de urgência
        resultado['nivel_urgencia'] = self._determinar_urgencia(
            resultado['diagnostico_consolidado']
        )
        
        # Gerar recomendação
        resultado['recomendacao'] = self._gerar_recomendacao(
            resultado['diagnostico_consolidado'],
            resultado['nivel_urgencia']
        )
        
        return resultado
    
    def _analisar_imagem_individual(self, img_path: str, numero: int) -> Dict:
        """
        Analisa uma imagem individual do pneu
        
        Nota: Esta é uma implementação simulada. Em produção, seria
        integrada com um modelo de visão computacional (OpenAI Vision,
        Google Vision API, ou modelo customizado)
        """
        # Simular análise de imagem
        # Em produção, aqui seria feita a chamada para API de visão computacional
        
        tipo_imagem = self._identificar_tipo_imagem(img_path)
        
        analise = {
            'numero': numero,
            'caminho': img_path,
            'tipo': tipo_imagem,
            'caracteristicas_detectadas': [],
            'problemas_identificados': [],
            'confianca': 0.0
        }
        
        # Baseado no tipo de imagem, simular detecções
        if tipo_imagem == 'banda_rodagem':
            analise['caracteristicas_detectadas'] = [
                'Desgaste irregular visível',
                'Profundidade de sulco reduzida',
                'Possível desgaste centralizado'
            ]
            analise['problemas_identificados'] = ['Desgaste Centralizado']
            analise['confianca'] = 0.85
            
        elif tipo_imagem == 'lateral':
            analise['caracteristicas_detectadas'] = [
                'Superfície lateral íntegra',
                'Sem bolhas ou rachaduras visíveis'
            ]
            analise['problemas_identificados'] = []
            analise['confianca'] = 0.90
            
        elif tipo_imagem == 'dot':
            analise['caracteristicas_detectadas'] = [
                'DOT legível',
                'Código de série identificado'
            ]
            analise['dot_detectado'] = '22412L18007'
            analise['confianca'] = 0.95
            
        elif tipo_imagem == 'defeito':
            analise['caracteristicas_detectadas'] = [
                'Dano estrutural visível',
                'Possível separação de cintas',
                'Área afetada: aproximadamente 15cm²'
            ]
            analise['problemas_identificados'] = [
                'Separação de Cintas',
                'Dano Estrutural'
            ]
            analise['confianca'] = 0.88
        
        return analise
    
    def _identificar_tipo_imagem(self, img_path: str) -> str:
        """
        Identifica o tipo de imagem baseado no nome do arquivo
        """
        nome_lower = os.path.basename(img_path).lower()
        
        if any(x in nome_lower for x in ['banda', 'rodagem', 'tread']):
            return 'banda_rodagem'
        elif any(x in nome_lower for x in ['lateral', 'sidewall', 'flanco']):
            return 'lateral'
        elif 'dot' in nome_lower:
            return 'dot'
        elif any(x in nome_lower for x in ['defeito', 'dano', 'damage']):
            return 'defeito'
        elif any(x in nome_lower for x in ['profundidade', 'sulco', 'depth']):
            return 'profundidade'
        else:
            return 'geral'
    
    def _consolidar_diagnostico(self, analises: List[Dict], contexto: Dict) -> Dict:
        """
        Consolida as análises individuais em um diagnóstico único
        """
        todos_problemas = []
        for analise in analises:
            todos_problemas.extend(analise.get('problemas_identificados', []))
        
        # Remover duplicatas
        problemas_unicos = list(set(todos_problemas))
        
        # Buscar informações detalhadas no banco de dados
        diagnostico = {
            'problemas_principais': problemas_unicos,
            'descricao_tecnica': '',
            'causas_provaveis': [],
            'defeito_fabricacao': False,
            'defeito_uso': False
        }
        
        # Usar IA para análise detalhada
        if contexto.get('descricao_usuario'):
            analise_ia = self.ia.diagnosticar_por_imagem(contexto['descricao_usuario'])
            diagnostico['analise_ia'] = analise_ia
        
        # Classificar tipo de defeito
        if any('Separação' in p or 'Fabricação' in p for p in problemas_unicos):
            diagnostico['defeito_fabricacao'] = True
            diagnostico['descricao_tecnica'] = 'Possível defeito de fabricação - Separação de componentes estruturais'
        else:
            diagnostico['defeito_uso'] = True
            diagnostico['descricao_tecnica'] = 'Desgaste relacionado a condições de uso'
        
        return diagnostico
    
    def _determinar_urgencia(self, diagnostico: Dict) -> str:
        """Determina o nível de urgência baseado no diagnóstico"""
        if diagnostico.get('defeito_fabricacao'):
            return 'ALTA'
        elif len(diagnostico.get('problemas_principais', [])) > 2:
            return 'MEDIA'
        else:
            return 'BAIXA'
    
    def _gerar_recomendacao(self, diagnostico: Dict, urgencia: str) -> str:
        """Gera recomendação baseada no diagnóstico"""
        if urgencia == 'ALTA':
            return 'Pneu deve ser retirado de operação imediatamente. Análise técnica presencial recomendada para avaliação de garantia.'
        elif urgencia == 'MEDIA':
            return 'Recomenda-se inspeção técnica detalhada. Possível elegibilidade para garantia dependendo de análise completa.'
        else:
            return 'Monitoramento recomendado. Documentar evolução do desgaste.'
    
    def _set_cell_value(self, ws, row: int, col: int, value):
        """
        Define valor de célula, lidando com células mescladas
        """
        try:
            cell = ws.cell(row, col)
            if hasattr(cell, '_value'):
                # Célula normal
                cell.value = value
            else:
                # Célula mesclada - tentar encontrar célula mestre
                for merged_range in ws.merged_cells.ranges:
                    if (row, col) in merged_range:
                        # Usar a célula superior esquerda da área mesclada
                        master_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                        master_cell.value = value
                        return
                # Se não encontrou mesclagem, tentar atribuir direto
                ws.cell(row, col).value = value
        except:
            # Silenciosamente ignorar erros de células mescladas
            pass
    
    def preencher_formulario_garantia(self, 
                                     analise_imagens: Dict,
                                     dados_cliente: Dict,
                                     dados_produto: Dict,
                                     output_path: str) -> str:
        """
        Preenche o formulário de garantia automaticamente
        
        Args:
            analise_imagens: Resultado da análise de imagens
            dados_cliente: Dados do consumidor/distribuidor
            dados_produto: Dados do pneu
            output_path: Caminho para salvar o formulário preenchido
        
        Returns:
            Caminho do arquivo gerado
        """
        # Carregar template
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        
        # Preencher Seção A - Consumidor
        if 'consumidor' in dados_cliente:
            self._preencher_secao_consumidor(ws, dados_cliente['consumidor'])
        
        # Preencher Seção B - Distribuidor
        if 'distribuidor' in dados_cliente:
            self._preencher_secao_distribuidor(ws, dados_cliente['distribuidor'])
        
        # Preencher Seção C - Produto Reclamado
        self._preencher_secao_produto(ws, dados_produto)
        
        # Preencher Seção F - Defeito Alegado
        self._preencher_defeito_alegado(ws, analise_imagens)
        
        # Preencher Seção G - Resultado da Análise
        self._preencher_resultado_analise(ws, analise_imagens)
        
        # Adicionar timestamp
        ws['B2'] = f"Gerado automaticamente em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        # Salvar arquivo
        wb.save(output_path)
        wb.close()
        
        return output_path
    
    def _preencher_secao_consumidor(self, ws, dados: Dict):
        """Preenche dados do consumidor"""
        linha_base = 3
        
        try:
            # Nome do consumidor
            self._set_cell_value(ws, linha_base, 3, dados.get('nome', ''))
            
            # Endereço
            self._set_cell_value(ws, linha_base + 1, 3, dados.get('endereco', ''))
            
            # CEP
            if 'cep' in dados:
                cep = dados['cep'].replace('-', '')
                for i, digito in enumerate(cep):
                    self._set_cell_value(ws, linha_base + 1, 22 + i, digito)
            
            # Cidade e UF
            self._set_cell_value(ws, linha_base + 2, 3, dados.get('cidade', ''))
            self._set_cell_value(ws, linha_base + 2, 20, dados.get('uf', ''))
            
            # Telefone
            self._set_cell_value(ws, linha_base + 2, 23, dados.get('telefone', ''))
        except Exception as e:
            print(f"Aviso ao preencher consumidor: {e}")
    
    def _preencher_secao_distribuidor(self, ws, dados: Dict):
        """Preenche dados do distribuidor"""
        linha_base = 7
        
        try:
            self._set_cell_value(ws, linha_base, 15, dados.get('nome', ''))
            self._set_cell_value(ws, linha_base + 1, 4, dados.get('endereco', ''))
            self._set_cell_value(ws, linha_base + 2, 4, dados.get('cidade', ''))
            self._set_cell_value(ws, linha_base + 2, 20, dados.get('uf', ''))
            self._set_cell_value(ws, linha_base + 2, 24, dados.get('telefone', ''))
            self._set_cell_value(ws, linha_base + 3, 3, dados.get('cnpj', ''))
        except Exception as e:
            print(f"Aviso ao preencher distribuidor: {e}")
    
    def _preencher_secao_produto(self, ws, dados: Dict):
        """Preenche dados do produto"""
        linha_base = 13
        
        try:
            # Medida
            self._set_cell_value(ws, linha_base, 8, dados.get('medida', ''))
            
            # Modelo
            self._set_cell_value(ws, linha_base, 11, dados.get('modelo', ''))
            
            # Marca
            self._set_cell_value(ws, linha_base, 14, dados.get('marca', ''))
            
            # Lonas/Índice de carga
            self._set_cell_value(ws, linha_base, 17, dados.get('lonas', ''))
            
            # Série e DOT
            self._set_cell_value(ws, linha_base, 22, dados.get('serie_dot', ''))
            
            # KM do produto
            self._set_cell_value(ws, linha_base, 31, dados.get('km', '0 km'))
        except Exception as e:
            print(f"Aviso ao preencher produto: {e}")
    
    def _preencher_defeito_alegado(self, ws, analise: Dict):
        """Preenche a descrição do defeito"""
        linha_base = 22
        
        try:
            diagnostico = analise.get('diagnostico_consolidado', {})
            descricao = diagnostico.get('descricao_tecnica', '')
            
            # Adicionar problemas identificados
            problemas = diagnostico.get('problemas_principais', [])
            if problemas:
                descricao += f"\n\nProblemas identificados pela IA: {', '.join(problemas)}"
            
            # Adicionar recomendação
            descricao += f"\n\nRecomendação: {analise.get('recomendacao', '')}"
            
            self._set_cell_value(ws, linha_base, 2, descricao)
        except Exception as e:
            print(f"Aviso ao preencher defeito alegado: {e}")
    
    def _preencher_resultado_analise(self, ws, analise: Dict):
        """Preenche o resultado da análise técnica"""
        linha_base = 28
        
        try:
            # Adicionar resumo da análise de IA
            resumo = f"Análise realizada por IA em {analise.get('timestamp', '')}\n"
            resumo += f"Total de imagens analisadas: {analise.get('total_imagens', 0)}\n"
            resumo += f"Nível de urgência: {analise.get('nivel_urgencia', 'N/A')}\n\n"
            
            diagnostico = analise.get('diagnostico_consolidado', {})
            if diagnostico.get('defeito_fabricacao'):
                resumo += "POSSÍVEL DEFEITO DE FABRICAÇÃO\n"
            
            resumo += f"\n{diagnostico.get('descricao_tecnica', '')}"
            
            self._set_cell_value(ws, linha_base + 2, 2, resumo)
        except Exception as e:
            print(f"Aviso ao preencher resultado análise: {e}")
    
    def gerar_relatorio_completo(self,
                                 analise_imagens: Dict,
                                 dados_completos: Dict,
                                 output_dir: str) -> Dict:
        """
        Gera relatório completo incluindo formulário e documentação
        
        Returns:
            Dicionário com caminhos dos arquivos gerados
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Gerar formulário Excel
        formulario_path = os.path.join(
            output_dir,
            f'formulario_garantia_{timestamp}.xlsx'
        )
        
        self.preencher_formulario_garantia(
            analise_imagens,
            dados_completos.get('cliente', {}),
            dados_completos.get('produto', {}),
            formulario_path
        )
        
        # Gerar relatório JSON
        relatorio_json_path = os.path.join(
            output_dir,
            f'relatorio_analise_{timestamp}.json'
        )
        
        with open(relatorio_json_path, 'w', encoding='utf-8') as f:
            json.dump({
                'analise_imagens': analise_imagens,
                'dados_completos': dados_completos,
                'timestamp': datetime.now().isoformat()
            }, f, indent=2, ensure_ascii=False)
        
        return {
            'formulario_excel': formulario_path,
            'relatorio_json': relatorio_json_path,
            'timestamp': timestamp
        }

def exemplo_uso():
    """Exemplo de uso do sistema"""
    print("="*70)
    print("SISTEMA DE ANÁLISE DE IMAGENS E PREENCHIMENTO DE GARANTIA")
    print("="*70 + "\n")
    
    # Inicializar analisador
    analisador = AnalisadorImagemGarantia()
    
    # Simular recebimento de imagens via WhatsApp
    print("📱 SIMULAÇÃO: Imagens recebidas via WhatsApp")
    print("-" * 70)
    
    imagens_simuladas = [
        '/path/to/imagem_banda_rodagem.jpg',
        '/path/to/imagem_lateral.jpg',
        '/path/to/imagem_dot.jpg',
        '/path/to/imagem_defeito.jpg'
    ]
    
    contexto = {
        'numero_whatsapp': '+55 47 99999-9999',
        'descricao_usuario': 'Pneu apresenta desgaste no centro e possível separação de cintas',
        'data_recebimento': datetime.now().isoformat()
    }
    
    # Analisar imagens
    print("🔍 Analisando imagens com IA...")
    analise = analisador.analisar_imagens_whatsapp(imagens_simuladas, contexto)
    
    print(f"\n✓ Análise concluída!")
    print(f"  • Total de imagens: {analise['total_imagens']}")
    print(f"  • Nível de urgência: {analise['nivel_urgencia']}")
    print(f"  • Problemas identificados: {len(analise['diagnostico_consolidado'].get('problemas_principais', []))}")
    
    # Dados para preenchimento do formulário
    print("\n\n📋 Preparando formulário de garantia...")
    print("-" * 70)
    
    dados_completos = {
        'cliente': {
            'consumidor': {
                'nome': 'Transportadora XYZ Ltda',
                'endereco': 'Rua das Flores, 123',
                'cep': '88000-000',
                'cidade': 'Florianópolis',
                'uf': 'SC',
                'telefone': '(48) 3333-4444'
            },
            'distribuidor': {
                'nome': 'PEDRONI LOGÍSTICA LTDA',
                'endereco': 'BR 101 Km 120, N° 8025',
                'cidade': 'Itajaí',
                'uf': 'SC',
                'telefone': '47 3348-8214',
                'cnpj': '01.832.166/0001-37'
            }
        },
        'produto': {
            'medida': '295/80R22.5',
            'modelo': 'Forza Plus',
            'marca': 'XBRI',
            'lonas': '18 lonas',
            'serie_dot': '22412L18007',
            'km': '0 km'
        }
    }
    
    # Gerar relatório completo
    output_dir = '/home/ubuntu/ia_analise_pneus/relatorios'
    os.makedirs(output_dir, exist_ok=True)
    
    arquivos = analisador.gerar_relatorio_completo(
        analise,
        dados_completos,
        output_dir
    )
    
    print(f"\n✓ Relatório gerado com sucesso!")
    print(f"  • Formulário Excel: {os.path.basename(arquivos['formulario_excel'])}")
    print(f"  • Relatório JSON: {os.path.basename(arquivos['relatorio_json'])}")
    
    print("\n" + "="*70)
    print("✅ Sistema funcionando corretamente!")
    print("="*70)

if __name__ == "__main__":
    exemplo_uso()

